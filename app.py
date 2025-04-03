from flask import Flask, request, jsonify, send_from_directory, send_file, g
from flask_cors import CORS
from datetime import datetime, timezone, timedelta
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors
import os
import qrcode
import base64
from io import BytesIO
import tempfile
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side
import sqlite3
import uuid
from werkzeug.security import generate_password_hash, check_password_hash
import jwt

app = Flask(__name__)
CORS(app)
app.config['DATABASE'] = 'database.db'
app.config['SECRET_KEY'] = 'e0fab1400bb09065526ef81315364fc81a6a0c05d77d47d30b009810be9c304a'  # Substitua por uma chave segura em produção

# Função para conectar ao banco de dados
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(app.config['DATABASE'])
        db.row_factory = sqlite3.Row
    return db

# Fechar conexão do banco ao finalizar o contexto
@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

# Inicializar o banco de dados
def init_db():
    with app.app_context():
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS trabalhadores (
                id TEXT PRIMARY KEY,
                nome TEXT NOT NULL,
                chefe BOOLEAN NOT NULL,
                qr_code_trabalhador TEXT,
                qr_code_chefe TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS paletes (
                id TEXT PRIMARY KEY,
                data_entrega TEXT,
                op TEXT,
                referencia TEXT,
                nome_produto TEXT,
                medida TEXT,
                cor_botao TEXT,
                cor_ribete TEXT,
                leva_embalagem BOOLEAN,
                quantidade INTEGER,
                data_hora TEXT,
                numero_lote TEXT,
                qr_code TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS tarefas (
                id TEXT PRIMARY KEY,
                nome TEXT NOT NULL,
                secao TEXT NOT NULL,
                qr_code TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS registros_trabalho (
                id TEXT PRIMARY KEY,
                data TEXT,
                palete_id TEXT,
                secao TEXT,
                tarefa_id TEXT,
                trabalhador_id TEXT,
                hora_inicio TEXT,
                hora_fim TEXT,
                FOREIGN KEY(palete_id) REFERENCES paletes(id),
                FOREIGN KEY(tarefa_id) REFERENCES tarefas(id),
                FOREIGN KEY(trabalhador_id) REFERENCES trabalhadores(id)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id TEXT PRIMARY KEY,
                email TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                tipo_usuario TEXT NOT NULL CHECK(tipo_usuario IN ('admin', 'chefe', 'funcionario'))
            )
        ''')
        
        db.commit()

init_db()

# Servir o arquivo HTML principal
@app.route('/')
def home():
    try:
        return send_from_directory(os.path.dirname(os.path.abspath(__file__)), 'index.html')
    except Exception as e:
        print(f"Erro ao servir index.html: {e}")
        return jsonify({'message': 'Erro ao servir a página inicial'}), 500

# Evitar erros de favicon
@app.route('/favicon.ico')
def no_favicon():
    return '', 204

@app.route('/favicon.png')
def no_favicon_png():
    return '', 204

# Função para gerar QR Code em base64
def gerar_qr_code_base64(conteudo):
    try:
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(conteudo)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buffered = BytesIO()
        img.save(buffered, format="PNG")
        return base64.b64encode(buffered.getvalue()).decode('utf-8')
    except Exception as e:
        print(f"Erro ao gerar QR Code: {e}")
        raise

# Decorador para exigir token JWT
def token_required(f):
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'message': 'Token não fornecido'}), 401
        try:
            token = token.split(" ")[1]  # Remove "Bearer"
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"])
            g.user = data
        except Exception as e:
            return jsonify({'message': 'Token inválido ou expirado', 'details': str(e)}), 401
        return f(*args, **kwargs)
    decorated.__name__ = f.__name__
    return decorated

# Registro de usuários (apenas admin inicial ou por admin autenticado)
@app.route('/register', methods=['POST'])
def register():
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        tipo_usuario = data.get('tipo_usuario')
        
        if not email or not password or not tipo_usuario:
            return jsonify({'message': 'Email, senha e tipo de usuário são obrigatórios'}), 400
        
        if tipo_usuario not in ['admin', 'chefe', 'funcionario']:
            return jsonify({'message': 'Tipo de usuário inválido. Use: admin, chefe ou funcionario'}), 400
        
        # Verificar se é o primeiro admin ou se o registro é feito por um admin autenticado
        token = request.headers.get('Authorization')
        if token:
            token = token.split(" ")[1]
            user_data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"])
            if user_data['tipo_usuario'] != 'admin':
                return jsonify({'message': 'Apenas admin pode registrar novos usuários'}), 403
            if tipo_usuario == 'admin':
                return jsonify({'message': 'Apenas um admin inicial pode ser registrado sem token'}), 403
        else:
            # Registro inicial do admin sem token
            cursor = get_db().cursor()
            cursor.execute('SELECT COUNT(*) as count FROM users WHERE tipo_usuario = "admin"')
            admin_count = cursor.fetchone()['count']
            if admin_count > 0 and tipo_usuario == 'admin':
                return jsonify({'message': 'Já existe um admin registrado. Use um token de admin para criar novos usuários'}), 403
        
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM users WHERE email = ?', (email,))
        if cursor.fetchone():
            return jsonify({'message': 'Email já cadastrado'}), 409
        
        user_id = str(uuid.uuid4())
        password_hash = generate_password_hash(password)
        
        cursor.execute('''
            INSERT INTO users (id, email, password_hash, tipo_usuario)
            VALUES (?, ?, ?, ?)
        ''', (user_id, email, password_hash, tipo_usuario))
        get_db().commit()
        
        token = jwt.encode({
            'id': user_id,
            'email': email,
            'tipo_usuario': tipo_usuario,
            'exp': datetime.now(timezone.utc) + timedelta(hours=24)
        }, app.config['SECRET_KEY'], algorithm='HS256')
        
        return jsonify({
            'message': 'Usuário registrado com sucesso!',
            'uid': user_id,
            'email': email,
            'tipo_usuario': tipo_usuario,
            'token': token
        }), 201
    except Exception as e:
        return jsonify({'message': 'Erro ao registrar usuário', 'details': str(e)}), 500

# Login de usuários
@app.route('/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        
        if not email or not password:
            return jsonify({'message': 'Email e senha são obrigatórios'}), 400
        
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM users WHERE email = ?', (email,))
        user = cursor.fetchone()
        
        if not user or not check_password_hash(user['password_hash'], password):
            return jsonify({'message': 'Credenciais inválidas'}), 401
        
        token = jwt.encode({
            'id': user['id'],
            'email': user['email'],
            'tipo_usuario': user['tipo_usuario'],
            'exp': datetime.now(timezone.utc) + timedelta(hours=24)
        }, app.config['SECRET_KEY'], algorithm='HS256')
        
        return jsonify({
            'message': 'Login bem-sucedido',
            'uid': user['id'],
            'email': user['email'],
            'tipo_usuario': user['tipo_usuario'],
            'token': token
        }), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao fazer login', 'details': str(e)}), 500

# Verificar token
@app.route('/verify-token', methods=['POST'])
@token_required
def verify_token():
    try:
        return jsonify({
            'valid': True,
            'tipo_usuario': g.user['tipo_usuario']
        }), 200
    except Exception as e:
        return jsonify({'valid': False, 'message': 'Erro ao verificar token', 'details': str(e)}), 500

# Listar usuários (apenas admin)
@app.route('/users', methods=['GET'])
@token_required
def get_users():
    try:
        if g.user['tipo_usuario'] != 'admin':
            return jsonify({'message': 'Apenas admin pode listar usuários'}), 403
        
        cursor = get_db().cursor()
        cursor.execute('SELECT id, email, tipo_usuario FROM users')
        users = [dict(row) for row in cursor.fetchall()]
        return jsonify(users), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao listar usuários', 'details': str(e)}), 500

# Eliminar usuários (apenas admin) 
@app.route('/users/<string:user_id>', methods=['DELETE'])
@token_required
def delete_user(user_id):
    try:
        if g.user['tipo_usuario'] != 'admin':
            return jsonify({'message': 'Apenas admin pode deletar usuários'}), 403
        
        if g.user['id'] == user_id:
            return jsonify({'message': 'Você não pode deletar a si mesmo'}), 403
        
        cursor = get_db().cursor()
        cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
        get_db().commit()
        
        if cursor.rowcount == 0:
            return jsonify({'message': 'Usuário não encontrado'}), 404
        
        return jsonify({'message': 'Usuário deletado com sucesso!'}), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao deletar usuário', 'details': str(e)}), 500

# Listar trabalhadores
@app.route('/trabalhadores', methods=['GET'])
@token_required
def get_trabalhadores():
    try:
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM trabalhadores')
        trabalhadores = [dict(row) for row in cursor.fetchall()]
        return jsonify(trabalhadores), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao listar trabalhadores', 'details': str(e)}), 500

# Adicionar trabalhador
@app.route('/trabalhadores', methods=['POST'])
@token_required
def add_trabalhador():
    try:
        if g.user['tipo_usuario'] not in ['admin', 'chefe']:
            return jsonify({'message': 'Apenas admin ou chefe pode adicionar trabalhadores'}), 403
        
        data = request.get_json()
        nome = data.get('nome')
        is_chefe = data.get('chefe', False)
        
        if not nome:
            return jsonify({'message': 'Nome do trabalhador é obrigatório'}), 400
        
        trabalhador_id = str(uuid.uuid4())
        qr_code_trabalhador_data = f"ID:{trabalhador_id};Tipo:Trabalhador;Nome:{nome}"
        qr_code_trabalhador = gerar_qr_code_base64(qr_code_trabalhador_data)
        
        qr_code_chefe = None
        if is_chefe:
            qr_code_chefe_data = f"ID:{trabalhador_id};Tipo:Chefe;Nome:{nome}"
            qr_code_chefe = gerar_qr_code_base64(qr_code_chefe_data)
        
        cursor = get_db().cursor()
        cursor.execute('''
            INSERT INTO trabalhadores (id, nome, chefe, qr_code_trabalhador, qr_code_chefe)
            VALUES (?, ?, ?, ?, ?)
        ''', (trabalhador_id, nome, is_chefe, qr_code_trabalhador, qr_code_chefe))
        get_db().commit()
        
        return jsonify({
            'message': 'Trabalhador adicionado com sucesso!',
            'id': trabalhador_id,
            'qr_code_trabalhador': qr_code_trabalhador,
            'qr_code_chefe': qr_code_chefe
        }), 201
    except Exception as e:
        return jsonify({'message': 'Erro ao adicionar trabalhador', 'details': str(e)}), 500

# Deletar trabalhador
@app.route('/trabalhadores/<string:id>', methods=['DELETE'])
@token_required
def delete_trabalhador(id):
    try:
        if g.user['tipo_usuario'] not in ['admin', 'chefe']:
            return jsonify({'message': 'Apenas admin ou chefe pode deletar trabalhadores'}), 403
        
        cursor = get_db().cursor()
        cursor.execute('DELETE FROM trabalhadores WHERE id = ?', (id,))
        get_db().commit()
        
        if cursor.rowcount == 0:
            return jsonify({'message': 'Trabalhador não encontrado'}), 404
        
        return jsonify({'message': 'Trabalhador removido com sucesso!'}), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao remover trabalhador', 'details': str(e)}), 500

# Gerar cartão do trabalhador
@app.route('/cartao/<string:trabalhador_id>/<string:tipo_cartao>', methods=['GET'])
@token_required
def gerar_cartao_pdf(trabalhador_id, tipo_cartao):
    try:
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM trabalhadores WHERE id = ?', (trabalhador_id,))
        trabalhador = cursor.fetchone()
        
        if not trabalhador:
            return jsonify({'message': 'Trabalhador não encontrado'}), 404
        
        if tipo_cartao not in ['trabalhador', 'chefe']:
            return jsonify({'message': 'Tipo de cartão inválido. Use: trabalhador ou chefe'}), 400
        
        if tipo_cartao == 'chefe' and not trabalhador['chefe']:
            return jsonify({'message': 'Este trabalhador não é chefe'}), 400
        
        qr_code_base64 = trabalhador['qr_code_chefe'] if tipo_cartao == 'chefe' else trabalhador['qr_code_trabalhador']
        cor_cartao = colors.red if tipo_cartao == 'chefe' else colors.blue
        titulo_cartao = "Cartão de Chefe" if tipo_cartao == 'chefe' else "Cartão de Trabalhador"
        
        qr_code_data = base64.b64decode(qr_code_base64)
        qr_code_img = Image.open(BytesIO(qr_code_data))
        
        temp_dir = tempfile.gettempdir()
        pdf_path = os.path.join(temp_dir, f"cartao_{trabalhador_id}_{tipo_cartao}.pdf")
        largura, altura = 7 * cm, 10 * cm
        
        pdf = canvas.Canvas(pdf_path, pagesize=(largura, altura))
        pdf.setFillColor(cor_cartao)
        pdf.rect(0, 0, largura, altura, fill=True, stroke=False)
        
        qr_code_img = qr_code_img.resize((100, 100))
        qr_code_buffer = BytesIO()
        qr_code_img.save(qr_code_buffer, format="PNG")
        pdf.drawImage(ImageReader(qr_code_buffer), (largura - 100) / 2, altura - 120, 100, 100)
        
        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(20, altura - 150, f"Nome: {trabalhador['nome']}")
        pdf.drawString(20, altura - 170, f"ID: {trabalhador_id}")
        pdf.drawString(20, altura - 190, titulo_cartao)
        
        pdf.save()
        return send_file(pdf_path, as_attachment=True, download_name=f"cartao_{trabalhador_id}_{tipo_cartao}.pdf")
    except Exception as e:
        return jsonify({'message': 'Erro ao gerar cartão', 'details': str(e)}), 500

# Listar paletes
@app.route('/paletes', methods=['GET'])
@token_required
def get_paletes():
    try:
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM paletes')
        paletes = [dict(row) for row in cursor.fetchall()]
        return jsonify(paletes), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao listar paletes', 'details': str(e)}), 500

# Adicionar palete
@app.route('/paletes', methods=['POST'])
@token_required
def add_palete():
    try:
        if g.user['tipo_usuario'] not in ['admin', 'chefe']:
            return jsonify({'message': 'Apenas admin ou chefe pode adicionar paletes'}), 403
        
        data = request.get_json()
        required_fields = ['data_entrega', 'op', 'referencia', 'nome_produto', 'medida', 
                          'cor_botao', 'cor_ribete', 'leva_embalagem', 'quantidade', 
                          'data_hora', 'numero_lote']
        
        for field in required_fields:
            if field not in data:
                return jsonify({'message': f'Campo obrigatório ausente: {field}'}), 400
        
        palete_id = str(uuid.uuid4())
        conteudo_qr = f"ID:{palete_id};Referencia:{data['referencia']};Nome:{data['nome_produto']};NumeroLote:{data['numero_lote']}"
        qr_code = gerar_qr_code_base64(conteudo_qr)
        
        cursor = get_db().cursor()
        cursor.execute('''
            INSERT INTO paletes (
                id, data_entrega, op, referencia, nome_produto, medida, 
                cor_botao, cor_ribete, leva_embalagem, quantidade, data_hora, 
                numero_lote, qr_code
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            palete_id, data['data_entrega'], data['op'], data['referencia'],
            data['nome_produto'], data['medida'], data['cor_botao'],
            data['cor_ribete'], data['leva_embalagem'], data['quantidade'],
            data['data_hora'], data['numero_lote'], qr_code
        ))
        get_db().commit()
        
        return jsonify({
            'message': 'Palete adicionada com sucesso!',
            'palete_id': palete_id,
            'qr_code': qr_code
        }), 201
    except Exception as e:
        return jsonify({'message': 'Erro ao adicionar palete', 'details': str(e)}), 500

# Deletar palete
@app.route('/paletes/<string:palete_id>', methods=['DELETE'])
@token_required
def delete_palete(palete_id):
    try:
        if g.user['tipo_usuario'] not in ['admin', 'chefe']:
            return jsonify({'message': 'Apenas admin ou chefe pode deletar paletes'}), 403
        
        cursor = get_db().cursor()
        cursor.execute('DELETE FROM paletes WHERE id = ?', (palete_id,))
        get_db().commit()
        
        if cursor.rowcount == 0:
            return jsonify({'message': 'Palete não encontrada'}), 404
        
        return jsonify({'message': 'Palete removida com sucesso!'}), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao remover palete', 'details': str(e)}), 500

# Gerar PDF da palete
@app.route('/paletes/<string:palete_id>/pdf', methods=['GET'])
@token_required
def gerar_pdf_palete(palete_id):
    try:
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM paletes WHERE id = ?', (palete_id,))
        palete = cursor.fetchone()
        
        if not palete:
            return jsonify({'message': 'Palete não encontrada'}), 404
        
        qr_code_data = base64.b64decode(palete['qr_code'])
        qr_code_img = Image.open(BytesIO(qr_code_data))
        
        temp_dir = tempfile.gettempdir()
        pdf_path = os.path.join(temp_dir, f"palete_{palete_id}.pdf")
        largura, altura = A4
        
        pdf = canvas.Canvas(pdf_path, pagesize=A4)
        pdf.setFillColor(colors.lightblue)
        pdf.rect(0, 0, largura, altura, fill=True, stroke=False)
        
        qr_code_size = 250
        qr_code_img = qr_code_img.resize((qr_code_size, qr_code_size))
        qr_code_buffer = BytesIO()
        qr_code_img.save(qr_code_buffer, format="PNG")
        pdf.drawImage(ImageReader(qr_code_buffer), (largura - qr_code_size) / 2, altura - 300, qr_code_size, qr_code_size)
        
        pdf.setFillColor(colors.black)
        pdf.setFont("Helvetica", 12)
        y = altura - 320
        line_height = 14
        
        campos = [
            f"Data de Entrega: {palete['data_entrega']}",
            f"OP: {palete['op']}",
            f"Referência: {palete['referencia']}",
            f"Nome do Produto: {palete['nome_produto']}",
            f"Medida: {palete['medida']}",
            f"Cor do Botão: {palete['cor_botao']}",
            f"Cor do Ribete: {palete['cor_ribete']}",
            f"Leva Embalagem: {'Sim' if palete['leva_embalagem'] else 'Não'}",
            f"Quantidade: {palete['quantidade']}",
            f"Data e Hora: {palete['data_hora']}",
            f"Número do Lote: {palete['numero_lote']}"
        ]
        
        for campo in campos:
            pdf.drawString(50, y, campo)
            y -= line_height
        
        pdf.save()
        return send_file(pdf_path, as_attachment=True, download_name=f"palete_{palete_id}.pdf")
    except Exception as e:
        return jsonify({'message': 'Erro ao gerar PDF', 'details': str(e)}), 500

# Listar tarefas
@app.route('/tarefas', methods=['GET'])
@token_required
def get_tarefas():
    try:
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM tarefas')
        tarefas = [dict(row) for row in cursor.fetchall()]
        return jsonify(tarefas), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao listar tarefas', 'details': str(e)}), 500

# Adicionar tarefa
@app.route('/tarefas', methods=['POST'])
@token_required
def add_tarefa():
    try:
        if g.user['tipo_usuario'] not in ['admin', 'chefe']:
            return jsonify({'message': 'Apenas admin ou chefe pode adicionar tarefas'}), 403
        
        data = request.get_json()
        nome_tarefa = data.get('nome_tarefa')
        secao = data.get('secao')
        
        if not nome_tarefa or not secao:
            return jsonify({'message': 'Nome da tarefa e seção são obrigatórios'}), 400
        
        tarefa_id = str(uuid.uuid4())
        qr_code_data = f"ID:{tarefa_id};Tarefa:{nome_tarefa};Secao:{secao}"
        qr_code = gerar_qr_code_base64(qr_code_data)
        
        cursor = get_db().cursor()
        cursor.execute('''
            INSERT INTO tarefas (id, nome, secao, qr_code)
            VALUES (?, ?, ?, ?)
        ''', (tarefa_id, nome_tarefa, secao, qr_code))
        get_db().commit()
        
        return jsonify({
            'message': 'Tarefa adicionada com sucesso!',
            'tarefa_id': tarefa_id,
            'qr_code': qr_code
        }), 201
    except Exception as e:
        return jsonify({'message': 'Erro ao adicionar tarefa', 'details': str(e)}), 500

# Deletar tarefa
@app.route('/tarefas/<string:tarefa_id>', methods=['DELETE'])
@token_required
def delete_tarefa(tarefa_id):
    try:
        if g.user['tipo_usuario'] not in ['admin', 'chefe']:
            return jsonify({'message': 'Apenas admin ou chefe pode deletar tarefas'}), 403
        
        cursor = get_db().cursor()
        cursor.execute('DELETE FROM tarefas WHERE id = ?', (tarefa_id,))
        get_db().commit()
        
        if cursor.rowcount == 0:
            return jsonify({'message': 'Tarefa não encontrada'}), 404
        
        return jsonify({'message': 'Tarefa removida com sucesso!'}), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao remover tarefa', 'details': str(e)}), 500

# Gerar PDF da tarefa
@app.route('/tarefas/<string:tarefa_id>/pdf', methods=['GET'])
@token_required
def gerar_pdf_tarefa(tarefa_id):
    try:
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM tarefas WHERE id = ?', (tarefa_id,))
        tarefa = cursor.fetchone()
        
        if not tarefa:
            return jsonify({'message': 'Tarefa não encontrada'}), 404
        
        qr_code_data = base64.b64decode(tarefa['qr_code'])
        qr_code_img = Image.open(BytesIO(qr_code_data))
        
        temp_dir = tempfile.gettempdir()
        pdf_path = os.path.join(temp_dir, f"tarefa_{tarefa_id}.pdf")
        largura, altura = (14 * cm, 21 * cm)
        
        pdf = canvas.Canvas(pdf_path, pagesize=(largura, altura))
        pdf.setFillColor(colors.lightgreen)
        pdf.rect(0, 0, largura, altura, fill=True, stroke=False)
        
        qr_code_size = 150
        qr_code_img = qr_code_img.resize((qr_code_size, qr_code_size))
        qr_code_buffer = BytesIO()
        qr_code_img.save(qr_code_buffer, format="PNG")
        pdf.drawImage(ImageReader(qr_code_buffer), (largura - qr_code_size) / 2, altura - 200, qr_code_size, qr_code_size)
        
        pdf.setFillColor(colors.black)
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(30, altura - 220, f"Tarefa: {tarefa['nome']}")
        pdf.drawString(30, altura - 240, f"Seção: {tarefa['secao']}")
        
        pdf.save()
        return send_file(pdf_path, as_attachment=True, download_name=f"tarefa_{tarefa_id}.pdf")
    except Exception as e:
        return jsonify({'message': 'Erro ao gerar PDF', 'details': str(e)}), 500

# Listar registros de trabalho
@app.route('/registro_trabalho', methods=['GET'])
@token_required
def get_registro_trabalho():
    try:
        cursor = get_db().cursor()
        cursor.execute('''
            SELECT rt.*, t.nome as tarefa_nome, tr.nome as trabalhador_nome, p.referencia
            FROM registros_trabalho rt
            JOIN tarefas t ON rt.tarefa_id = t.id
            JOIN trabalhadores tr ON rt.trabalhador_id = tr.id
            JOIN paletes p ON rt.palete_id = p.id
        ''')
        registros = []
        for row in cursor.fetchall():
            registros.append({
                'id': row['id'],
                'data': row['data'],
                'palete': {
                    'id': row['palete_id'],
                    'referencia': row['referencia']
                },
                'secao': row['secao'],
                'tarefa': {
                    'id': row['tarefa_id'],
                    'nome': row['tarefa_nome']
                },
                'trabalhador': {
                    'id': row['trabalhador_id'],
                    'nome': row['trabalhador_nome']
                },
                'hora_inicio': row['hora_inicio'],
                'hora_fim': row['hora_fim']
            })
        return jsonify(registros), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao listar registros', 'details': str(e)}), 500

# Registrar trabalho
@app.route('/registro_trabalho', methods=['POST'])
@token_required
def registrar_trabalho():
    try:
        if g.user['tipo_usuario'] not in ['chefe', 'funcionario']:
            return jsonify({'message': 'Apenas chefe ou funcionário pode registrar trabalho'}), 403
        
        data = request.get_json()
        tarefa_qr = data.get('tarefa_qr')
        trabalhador_qr = data.get('trabalhador_qr')
        palete_qr = data.get('palete_qr')
        
        if not all([tarefa_qr, trabalhador_qr, palete_qr]):
            return jsonify({'message': 'Todos os QR Codes são obrigatórios'}), 400
        
        tarefa_id = extrair_valor(tarefa_qr, "ID:")
        trabalhador_id = extrair_valor(trabalhador_qr, "ID:")
        palete_id = extrair_valor(palete_qr, "ID:")
        secao = extrair_valor(tarefa_qr, "Secao:") or "Default"
        
        cursor = get_db().cursor()
        
        cursor.execute('SELECT * FROM tarefas WHERE id = ?', (tarefa_id,))
        if not cursor.fetchone():
            return jsonify({'message': 'Tarefa não encontrada'}), 404
        
        cursor.execute('SELECT * FROM trabalhadores WHERE id = ?', (trabalhador_id,))
        if not cursor.fetchone():
            return jsonify({'message': 'Trabalhador não encontrado'}), 404
        
        cursor.execute('SELECT * FROM paletes WHERE id = ?', (palete_id,))
        if not cursor.fetchone():
            return jsonify({'message': 'Palete não encontrada'}), 404
        
        cursor.execute('''
            SELECT * FROM registros_trabalho 
            WHERE tarefa_id = ? AND palete_id = ? AND trabalhador_id = ? AND hora_fim IS NULL
        ''', (tarefa_id, palete_id, trabalhador_id))
        registro_existente = cursor.fetchone()
        
        if registro_existente:
            cursor.execute('''
                UPDATE registros_trabalho 
                SET hora_fim = ? 
                WHERE id = ?
            ''', (datetime.now(timezone.utc).isoformat(), registro_existente['id']))
            get_db().commit()
            return jsonify({'message': 'Tarefa finalizada com sucesso!'}), 200
        
        registro_id = str(uuid.uuid4())
        cursor.execute('''
            INSERT INTO registros_trabalho (
                id, data, palete_id, secao, tarefa_id, 
                trabalhador_id, hora_inicio
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            registro_id, datetime.now(timezone.utc).isoformat(), palete_id, secao, tarefa_id,
            trabalhador_id, datetime.now(timezone.utc).isoformat()
        ))
        get_db().commit()
        
        return jsonify({'message': 'Registro de trabalho iniciado com sucesso!'}), 201
    except Exception as e:
        return jsonify({'message': 'Erro ao registrar trabalho', 'details': str(e)}), 500

# Função auxiliar para extrair valores do QR Code
def extrair_valor(qr_code, chave):
    try:
        partes = qr_code.split(";")
        for parte in partes:
            if parte.startswith(chave):
                return parte[len(chave):]
        return None
    except Exception as e:
        print(f"Erro ao extrair valor do QR Code: {e}")
        return None

# Exportar registros para Excel
@app.route('/exportar_registros', methods=['GET'])
@token_required
def exportar_registros():
    try:
        if g.user['tipo_usuario'] not in ['admin', 'chefe']:
            return jsonify({'message': 'Apenas admin ou chefe pode exportar registros'}), 403
        
        cursor = get_db().cursor()
        cursor.execute('''
            SELECT rt.data, p.referencia, t.nome as tarefa_nome, tr.nome as trabalhador_nome,
                   rt.hora_inicio, rt.hora_fim
            FROM registros_trabalho rt
            JOIN paletes p ON rt.palete_id = p.id
            JOIN tarefas t ON rt.tarefa_id = t.id
            JOIN trabalhadores tr ON rt.trabalhador_id = tr.id
        ''')
        
        registros = [dict(row) for row in cursor.fetchall()]
        df = pd.DataFrame(registros)
        
        temp_dir = tempfile.gettempdir()
        excel_path = os.path.join(temp_dir, "registros_trabalho.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros de Trabalho"
        
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.font = Font(bold=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
        
        wb.save(excel_path)
        return send_file(excel_path, as_attachment=True, download_name="registros_trabalho.xlsx")
    except Exception as e:
        return jsonify({'message': 'Erro ao exportar registros', 'details': str(e)}), 500

# Iniciar o servidor
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000)