from flask import Flask, request, jsonify, send_from_directory, send_file, g
from flask_cors import CORS
from datetime import datetime, timezone, timedelta
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors
import re
import os
import json
import qrcode
import base64
from io import BytesIO
import tempfile
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side
import sqlite3
import uuid
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
CORS(app)
app.config['DATABASE'] = 'database.db'

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(app.config['DATABASE'])
        db.row_factory = sqlite3.Row
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    with app.app_context():
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS trabalhadores (
                id TEXT PRIMARY KEY,
                nome TEXT NOT NULL,
                chefe BOOLEAN NOT NULL,
                qr_code_trabalhador TEXT,
                qr_code_chefe TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS paletes (
                id TEXT PRIMARY KEY,
                data_entrega TEXT,
                op TEXT,
                referencia TEXT,
                nome_produto TEXT,
                medida TEXT,
                cor_botao TEXT,
                cor_ribete TEXT,
                leva_embalagem BOOLEAN,
                quantidade INTEGER,
                data_hora TEXT,
                numero_lote TEXT,
                qr_code TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS tarefas (
                id TEXT PRIMARY KEY,
                nome TEXT NOT NULL,
                secao TEXT NOT NULL,
                qr_code TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS registros_trabalho (
                id TEXT PRIMARY KEY,
                data TEXT,
                palete_id TEXT,
                secao TEXT,
                tarefa_id TEXT,
                trabalhador_id TEXT,
                hora_inicio TEXT,
                hora_fim TEXT,
                FOREIGN KEY(palete_id) REFERENCES paletes(id),
                FOREIGN KEY(tarefa_id) REFERENCES tarefas(id),
                FOREIGN KEY(trabalhador_id) REFERENCES trabalhadores(id)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id TEXT PRIMARY KEY,
                email TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL
            )
        ''')
        
        db.commit()

init_db()

@app.route('/')
def home():
    try:
        return send_from_directory(os.path.dirname(os.path.abspath(__file__)), 'index.html')
    except Exception as e:
        print(f"Erro ao servir index.html: {e}")
        return "Erro ao servir a página inicial.", 500

@app.route('/favicon.ico')
def no_favicon():
    return '', 204

@app.route('/favicon.png')
def no_favicon_png():
    return '', 204

def gerar_qr_code_base64(conteudo):
    try:
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(conteudo)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buffered = BytesIO()
        img.save(buffered, format="PNG")
        return base64.b64encode(buffered.getvalue()).decode('utf-8')
    except Exception as e:
        print(f"Erro ao gerar QR Code: {e}")
        raise

@app.route('/trabalhadores', methods=['GET'])
def get_trabalhadores():
    try:
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM trabalhadores')
        trabalhadores = [dict(row) for row in cursor.fetchall()]
        return jsonify(trabalhadores), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao listar trabalhadores.', 'details': str(e)}), 500

@app.route('/trabalhadores', methods=['POST'])
def add_trabalhador():
    try:
        data = request.get_json()
        nome = data.get('nome')
        is_chefe = data.get('chefe', False)
        
        if not nome:
            return jsonify({'message': 'Nome do trabalhador é obrigatório.'}), 400
        
        trabalhador_id = str(uuid.uuid4())
        qr_code_trabalhador_data = f"ID:{trabalhador_id};Tipo:Trabalhador;Nome:{nome}"
        qr_code_trabalhador = gerar_qr_code_base64(qr_code_trabalhador_data)
        
        qr_code_chefe = None
        if is_chefe:
            qr_code_chefe_data = f"ID:{trabalhador_id};Tipo:Chefe;Nome:{nome}"
            qr_code_chefe = gerar_qr_code_base64(qr_code_chefe_data)
        
        cursor = get_db().cursor()
        cursor.execute('''
            INSERT INTO trabalhadores (id, nome, chefe, qr_code_trabalhador, qr_code_chefe)
            VALUES (?, ?, ?, ?, ?)
        ''', (trabalhador_id, nome, is_chefe, qr_code_trabalhador, qr_code_chefe))
        
        get_db().commit()
        
        return jsonify({
            'message': 'Trabalhador adicionado com sucesso!',
            'id': trabalhador_id,
            'qr_code_trabalhador': qr_code_trabalhador,
            'qr_code_chefe': qr_code_chefe
        }), 201
    except Exception as e:
        return jsonify({'message': 'Erro ao adicionar trabalhador.', 'details': str(e)}), 500

@app.route('/trabalhadores/<string:id>', methods=['DELETE'])
def delete_trabalhador(id):
    try:
        cursor = get_db().cursor()
        cursor.execute('DELETE FROM trabalhadores WHERE id = ?', (id,))
        get_db().commit()
        
        if cursor.rowcount == 0:
            return jsonify({'message': 'Trabalhador não encontrado'}), 404
            
        return jsonify({'message': 'Trabalhador removido com sucesso!'}), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao remover trabalhador.', 'details': str(e)}), 500

@app.route('/cartao/<string:trabalhador_id>/<string:tipo_cartao>', methods=['GET'])
def gerar_cartao_pdf(trabalhador_id, tipo_cartao):
    try:
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM trabalhadores WHERE id = ?', (trabalhador_id,))
        trabalhador = cursor.fetchone()
        
        if not trabalhador:
            return jsonify({'message': 'Trabalhador não encontrado.'}), 404
            
        if tipo_cartao == "chefe" and trabalhador['chefe']:
            qr_code_base64 = trabalhador['qr_code_chefe']
            cor_cartao = colors.red
            titulo_cartao = "Cartão de Chefe"
        else:
            qr_code_base64 = trabalhador['qr_code_trabalhador']
            cor_cartao = colors.blue
            titulo_cartao = "Cartão de Trabalhador"
        
        if not qr_code_base64:
            return jsonify({'message': 'QR Code não encontrado para este tipo de cartão.'}), 404
        
        qr_code_data = base64.b64decode(qr_code_base64)
        qr_code_img = Image.open(BytesIO(qr_code_data))
        
        temp_dir = tempfile.gettempdir()
        pdf_path = f"{temp_dir}/cartao_{trabalhador_id}_{tipo_cartao}.pdf"
        largura, altura = 7 * cm, 10 * cm
        
        pdf = canvas.Canvas(pdf_path, pagesize=(largura, altura))
        pdf.setFillColor(cor_cartao)
        pdf.rect(0, 0, largura, altura, fill=True, stroke=False)
        
        qr_code_img = qr_code_img.resize((100, 100))
        qr_code_buffer = BytesIO()
        qr_code_img.save(qr_code_buffer, format="PNG")
        pdf.drawImage(ImageReader(qr_code_buffer), (largura - 100)/2, altura - 120, 100, 100)
        
        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(20, altura - 150, f"Nome: {trabalhador['nome']}")
        pdf.drawString(20, altura - 170, f"ID: {trabalhador_id}")
        pdf.drawString(20, altura - 190, titulo_cartao)
        
        pdf.save()
        return send_from_directory(temp_dir, f"cartao_{trabalhador_id}_{tipo_cartao}.pdf", as_attachment=True)
    except Exception as e:
        print(f"Erro ao gerar cartão: {e}")
        return jsonify({'message': 'Erro ao gerar cartão.', 'details': str(e)}), 500

@app.route('/paletes', methods=['GET'])
def get_paletes():
    try:
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM paletes')
        paletes = [dict(row) for row in cursor.fetchall()]
        return jsonify(paletes), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao listar paletes.', 'details': str(e)}), 500

@app.route('/paletes', methods=['POST'])
def add_palete():
    try:
        data = request.get_json()
        required_fields = ['data_entrega', 'op', 'referencia', 'nome_produto', 
                         'medida', 'cor_botao', 'cor_ribete', 'leva_embalagem',
                         'quantidade', 'data_hora', 'numero_lote']
        
        for field in required_fields:
            if field not in data:
                return jsonify({'message': f'Campo obrigatório ausente: {field}'}), 400
        
        palete_id = str(uuid.uuid4())
        conteudo_qr = f"ID:{palete_id};Referencia:{data['referencia']};Nome:{data['nome_produto']};NumeroLote:{data['numero_lote']}"
        qr_code = gerar_qr_code_base64(conteudo_qr)
        
        cursor = get_db().cursor()
        cursor.execute('''
            INSERT INTO paletes (
                id, data_entrega, op, referencia, nome_produto, medida, 
                cor_botao, cor_ribete, leva_embalagem, quantidade, data_hora, 
                numero_lote, qr_code
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            palete_id, data['data_entrega'], data['op'], data['referencia'],
            data['nome_produto'], data['medida'], data['cor_botao'],
            data['cor_ribete'], data['leva_embalagem'], data['quantidade'],
            data['data_hora'], data['numero_lote'], qr_code
        ))
        
        get_db().commit()
        
        return jsonify({
            'message': 'Palete adicionada com sucesso!',
            'palete_id': palete_id,
            'qr_code': qr_code
        }), 201
    except Exception as e:
        return jsonify({'message': 'Erro ao adicionar palete.', 'details': str(e)}), 500

@app.route('/paletes/<string:palete_id>', methods=['DELETE'])
def delete_palete(palete_id):
    try:
        cursor = get_db().cursor()
        cursor.execute('DELETE FROM paletes WHERE id = ?', (palete_id,))
        get_db().commit()
        
        if cursor.rowcount == 0:
            return jsonify({'message': 'Palete não encontrada'}), 404
            
        return jsonify({'message': 'Palete removida com sucesso!'}), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao remover palete.', 'details': str(e)}), 500

@app.route('/paletes/<string:palete_id>/pdf', methods=['GET'])
def gerar_pdf_palete(palete_id):
    try:
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM paletes WHERE id = ?', (palete_id,))
        palete = cursor.fetchone()
        
        if not palete:
            return jsonify({'message': 'Palete não encontrada.'}), 404
            
        qr_code_data = base64.b64decode(palete['qr_code'])
        qr_code_img = Image.open(BytesIO(qr_code_data))
        
        temp_dir = tempfile.gettempdir()
        pdf_path = os.path.join(temp_dir, f"palete_{palete_id}.pdf")
        largura, altura = A4
        
        pdf = canvas.Canvas(pdf_path, pagesize=A4)
        pdf.setFillColor(colors.lightblue)
        pdf.rect(0, 0, largura, altura, fill=True, stroke=False)
        
        qr_code_size = 250
        qr_code_img = qr_code_img.resize((qr_code_size, qr_code_size))
        qr_code_buffer = BytesIO()
        qr_code_img.save(qr_code_buffer, format="PNG")
        pdf.drawImage(ImageReader(qr_code_buffer), (largura - qr_code_size)/2, altura - 300, qr_code_size, qr_code_size)
        
        pdf.setFillColor(colors.black)
        pdf.setFont("Helvetica", 12)
        y = altura - 320
        line_height = 14
        
        campos = [
            f"Data de Entrega: {palete['data_entrega']}",
            f"OP: {palete['op']}",
            f"Referência: {palete['referencia']}",
            f"Nome do Produto: {palete['nome_produto']}",
            f"Medida: {palete['medida']}",
            f"Cor do Botão: {palete['cor_botao']}",
            f"Cor do Ribete: {palete['cor_ribete']}",
            f"Leva Embalagem: {'Sim' if palete['leva_embalagem'] else 'Não'}",
            f"Quantidade: {palete['quantidade']}",
            f"Data e Hora: {palete['data_hora']}",
            f"Número do Lote: {palete['numero_lote']}"
        ]
        
        for campo in campos:
            pdf.drawString(50, y, campo)
            y -= line_height
        
        pdf.save()
        return send_from_directory(temp_dir, f"palete_{palete_id}.pdf", as_attachment=True)
    except Exception as e:
        return jsonify({'message': 'Erro ao gerar PDF.', 'details': str(e)}), 500

@app.route('/tarefas', methods=['GET'])
def get_tarefas():
    try:
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM tarefas')
        tarefas = [dict(row) for row in cursor.fetchall()]
        return jsonify(tarefas), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao listar tarefas.', 'details': str(e)}), 500

@app.route('/tarefas', methods=['POST'])
def add_tarefa():
    try:
        data = request.get_json()
        nome_tarefa = data.get('nome_tarefa')
        secao = data.get('secao')
        
        if not nome_tarefa or not secao:
            return jsonify({'message': 'Nome da tarefa e secção são obrigatórios.'}), 400
        
        tarefa_id = str(uuid.uuid4())
        qr_code_data = f"ID:{tarefa_id};Tarefa:{nome_tarefa};Secao:{secao}"
        qr_code = gerar_qr_code_base64(qr_code_data)
        
        cursor = get_db().cursor()
        cursor.execute('''
            INSERT INTO tarefas (id, nome, secao, qr_code)
            VALUES (?, ?, ?, ?)
        ''', (tarefa_id, nome_tarefa, secao, qr_code))
        
        get_db().commit()
        
        return jsonify({
            'message': 'Tarefa adicionada com sucesso!',
            'tarefa_id': tarefa_id,
            'qr_code': qr_code
        }), 201
    except Exception as e:
        return jsonify({'message': 'Erro ao adicionar tarefa.', 'details': str(e)}), 500

@app.route('/tarefas/<string:tarefa_id>', methods=['DELETE'])
def delete_tarefa(tarefa_id):
    try:
        cursor = get_db().cursor()
        cursor.execute('DELETE FROM tarefas WHERE id = ?', (tarefa_id,))
        get_db().commit()
        
        if cursor.rowcount == 0:
            return jsonify({'message': 'Tarefa não encontrada'}), 404
            
        return jsonify({'message': 'Tarefa removida com sucesso!'}), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao remover tarefa.', 'details': str(e)}), 500

@app.route('/tarefas/<string:tarefa_id>/pdf', methods=['GET'])
def gerar_pdf_tarefa(tarefa_id):
    try:
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM tarefas WHERE id = ?', (tarefa_id,))
        tarefa = cursor.fetchone()
        
        if not tarefa:
            return jsonify({'message': 'Tarefa não encontrada.'}), 404
            
        qr_code_data = base64.b64decode(tarefa['qr_code'])
        qr_code_img = Image.open(BytesIO(qr_code_data))
        
        temp_dir = tempfile.gettempdir()
        pdf_path = os.path.join(temp_dir, f"tarefa_{tarefa_id}.pdf")
        largura, altura = (14*cm, 21*cm)
        
        pdf = canvas.Canvas(pdf_path, pagesize=(largura, altura))
        pdf.setFillColor(colors.lightgreen)
        pdf.rect(0, 0, largura, altura, fill=True, stroke=False)
        
        qr_code_size = 150
        qr_code_img = qr_code_img.resize((qr_code_size, qr_code_size))
        qr_code_buffer = BytesIO()
        qr_code_img.save(qr_code_buffer, format="PNG")
        pdf.drawImage(ImageReader(qr_code_buffer), (largura - qr_code_size)/2, altura - 200, qr_code_size, qr_code_size)
        
        pdf.setFillColor(colors.black)
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(30, altura - 220, f"Tarefa: {tarefa['nome']}")
        pdf.drawString(30, altura - 240, f"Secção: {tarefa['secao']}")
        
        pdf.save()
        return send_from_directory(temp_dir, f"tarefa_{tarefa_id}.pdf", as_attachment=True)
    except Exception as e:
        return jsonify({'message': 'Erro ao gerar PDF.', 'details': str(e)}), 500

@app.route('/registro_trabalho', methods=['GET'])
def get_registro_trabalho():
    try:
        cursor = get_db().cursor()
        cursor.execute('''
            SELECT rt.*, t.nome as tarefa_nome, tr.nome as trabalhador_nome, p.referencia
            FROM registros_trabalho rt
            JOIN tarefas t ON rt.tarefa_id = t.id
            JOIN trabalhadores tr ON rt.trabalhador_id = tr.id
            JOIN paletes p ON rt.palete_id = p.id
        ''')
        registros = []
        for row in cursor.fetchall():
            registros.append({
                'id': row['id'],
                'data': row['data'],
                'palete': {
                    'id': row['palete_id'],
                    'referencia': row['referencia']
                },
                'secao': row['secao'],
                'tarefa': {
                    'id': row['tarefa_id'],
                    'nome': row['tarefa_nome']
                },
                'trabalhador': {
                    'id': row['trabalhador_id'],
                    'nome': row['trabalhador_nome']
                },
                'hora_inicio': row['hora_inicio'],
                'hora_fim': row['hora_fim']
            })
        return jsonify(registros), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao listar registros.', 'details': str(e)}), 500

@app.route('/registro_trabalho', methods=['POST'])
def registrar_trabalho():
    try:
        data = request.get_json()
        tarefa_qr = data.get('tarefa_qr')
        trabalhador_qr = data.get('trabalhador_qr')
        palete_qr = data.get('palete_qr')
        
        if not all([tarefa_qr, trabalhador_qr, palete_qr]):
            return jsonify({'message': 'Todos os QR Codes são obrigatórios.'}), 400
        
        tarefa_id = extrair_valor(tarefa_qr, "ID:")
        trabalhador_id = extrair_valor(trabalhador_qr, "ID:")
        palete_id = extrair_valor(palete_qr, "ID:")
        secao = extrair_valor(tarefa_qr, "Secao:") or "Default"
        
        cursor = get_db().cursor()
        
        # Verificar existência da tarefa
        cursor.execute('SELECT * FROM tarefas WHERE id = ?', (tarefa_id,))
        if not cursor.fetchone():
            return jsonify({'message': 'Tarefa não encontrada'}), 404
        
        # Verificar existência do trabalhador
        cursor.execute('SELECT * FROM trabalhadores WHERE id = ?', (trabalhador_id,))
        if not cursor.fetchone():
            return jsonify({'message': 'Trabalhador não encontrado'}), 404
        
        # Verificar existência da palete
        cursor.execute('SELECT * FROM paletes WHERE id = ?', (palete_id,))
        if not cursor.fetchone():
            return jsonify({'message': 'Palete não encontrada'}), 404
        
        # Verificar se já existe registro aberto
        cursor.execute('''
            SELECT * FROM registros_trabalho 
            WHERE tarefa_id = ? AND palete_id = ? AND hora_fim IS NULL
        ''', (tarefa_id, palete_id))
        registro_existente = cursor.fetchone()
        
        if registro_existente:
            # Atualizar registro existente com hora_fim
            cursor.execute('''
                UPDATE registros_trabalho 
                SET hora_fim = ? 
                WHERE id = ?
            ''', (datetime.now().isoformat(), registro_existente['id']))
            get_db().commit()
            return jsonify({'message': 'Tarefa finalizada com sucesso!'}), 200
        
        # Criar novo registro
        registro_id = str(uuid.uuid4())
        cursor.execute('''
            INSERT INTO registros_trabalho (
                id, data, palete_id, secao, tarefa_id, 
                trabalhador_id, hora_inicio
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            registro_id, datetime.now().date().isoformat(), palete_id,
            secao, tarefa_id, trabalhador_id, datetime.now().isoformat()
        ))
        
        get_db().commit()
        return jsonify({'message': 'Registro iniciado com sucesso!', 'registro_id': registro_id}), 201
    except Exception as e:
        return jsonify({'message': 'Erro ao registrar trabalho.', 'details': str(e)}), 500

@app.route('/exportar_registros', methods=['GET'])
def exportar_registros():
    try:
        cursor = get_db().cursor()
        cursor.execute('''
            SELECT rt.data, p.referencia, rt.secao, t.nome as tarefa, 
                   tr.nome as trabalhador, rt.hora_inicio, rt.hora_fim
            FROM registros_trabalho rt
            JOIN paletes p ON rt.palete_id = p.id
            JOIN tarefas t ON rt.tarefa_id = t.id
            JOIN trabalhadores tr ON rt.trabalhador_id = tr.id
        ''')
        
        df = pd.DataFrame(cursor.fetchall(), columns=[
            'Data', 'Palete Referência', 'Secção', 'Tarefa', 
            'Trabalhador', 'Hora Início', 'Hora Fim'
        ])
        
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Registros')
        
        # Formatação do Excel
        workbook = writer.book
        worksheet = writer.sheets['Registros']
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        border = Border(left=Side(style='thin'), 
                       right=Side(style='thin'), 
                       top=Side(style='thin'), 
                       bottom=Side(style='thin'))
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
        
        writer.save()
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name='registros_trabalho.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'message': 'Erro ao exportar registros.', 'details': str(e)}), 500

@app.route('/register', methods=['POST'])
def register():
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        
        if not email or not password:
            return jsonify({'message': 'Email e senha são obrigatórios'}), 400
        
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM users WHERE email = ?', (email,))
        if cursor.fetchone():
            return jsonify({'message': 'Email já cadastrado'}), 409
        
        user_id = str(uuid.uuid4())
        password_hash = generate_password_hash(password)
        
        cursor.execute('''
            INSERT INTO users (id, email, password_hash)
            VALUES (?, ?, ?)
        ''', (user_id, email, password_hash))
        
        get_db().commit()
        
        return jsonify({
            'message': 'Usuário registrado com sucesso!',
            'uid': user_id,
            'email': email
        }), 201
    except Exception as e:
        return jsonify({'message': 'Erro interno no servidor', 'details': str(e)}), 500

@app.route('/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        
        cursor = get_db().cursor()
        cursor.execute('SELECT * FROM users WHERE email = ?', (email,))
        user = cursor.fetchone()
        
        if not user or not check_password_hash(user['password_hash'], password):
            return jsonify({'message': 'Credenciais inválidas'}), 401
        
        return jsonify({
            'message': 'Login bem-sucedido',
            'uid': user['id'],
            'email': user['email']
        }), 200
    except Exception as e:
        return jsonify({'message': 'Erro ao autenticar usuário.', 'details': str(e)}), 401

def extrair_valor(qr_text, chave):
    for item in qr_text.split(';'):
        if chave in item:
            return item.split(':', 1)[-1].strip()
    return None

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000)